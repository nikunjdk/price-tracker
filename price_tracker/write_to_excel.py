import logging
import os
import pandas as pd
from openpyxl.styles import PatternFill

from price_tracker.const import (
    COLUMN_PRODUCT_CURRENT_PRICE,
    COLUMN_PRODUCT_DELIVERY_LOCATIONS,
    COLUMN_PRODUCT_MINIMUM_PRICE,
    COLUMN_PRODUCT_QAUNTITY,
    COLUMN_PRODUCT_URL,
)


def write_to_excel(current_date_time: str, products: pd.DataFrame) -> None:
    log = logging.getLogger("price_tracker.write_to_excel")
    try:
        if not os.path.exists("results"):
            os.mkdir("results")
        # Define the output file path
        output_path = f"results/price_tracker_{current_date_time}.xlsx"

        products[COLUMN_PRODUCT_QAUNTITY] = products[COLUMN_PRODUCT_QAUNTITY].replace(
            -1, ""
        )
        products[COLUMN_PRODUCT_DELIVERY_LOCATIONS] = products[
            COLUMN_PRODUCT_DELIVERY_LOCATIONS
        ].apply(lambda x: ", ".join(x))

        # Write the DataFrame to an Excel file
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            products.to_excel(
                writer,
                sheet_name="Sheet1",
                index=False,
                header=[
                    COLUMN_PRODUCT_URL,
                    COLUMN_PRODUCT_CURRENT_PRICE,
                    COLUMN_PRODUCT_QAUNTITY,
                    COLUMN_PRODUCT_DELIVERY_LOCATIONS,
                    COLUMN_PRODUCT_MINIMUM_PRICE,
                ],
            )

            # Get the workbook and the worksheet
            worksheet = writer.sheets["Sheet1"]

            red_fill = PatternFill(
                start_color="FF0000", end_color="FF0000", fill_type="solid"
            )

            # Apply the fill color to the cells based on the condition
            for row in worksheet.iter_rows():
                (
                    _,
                    current_price_cell,
                    _,
                    _,
                    minimum_price_cell,
                ) = row
                if current_price_cell.value > minimum_price_cell.value:
                    minimum_price_cell.fill = red_fill

            # Save the changes to the Excel file
            # writer.close()
    except Exception as e:
        log.error(e)
        raise
